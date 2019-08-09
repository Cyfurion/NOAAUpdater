# -*- coding: utf-8 -*-
import ctypes
import pyodbc
import requests
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from passlib.hash import sha256_crypt
from tkinter import Button
from tkinter import Entry
from tkinter import Label
from tkinter import messagebox
from tkinter import Tk

MessageBox = ctypes.windll.user32.MessageBoxW
payload = {'product': 'CF6', 'station': 'NYC', 'recent': 'yes' }
cleared = False
server_user = ''
server_pass = ''

# Function definitions.
def auth():
    if (sha256_crypt.verify(root_field.get(), '$5$rounds=535000$LBgMfjzfpi3Oo44j$a.Wp0Hw42z8SmRQbhfmlUHUeEwsr/euG1E1OHm6jQhC')):
        global cleared
        cleared = True
        root.destroy()
    else:
        messagebox.showerror('Authentication Error', 'The password entered was incorrect. Please try again.')
        
def server_submit():
    global server_user
    global server_pass
    server_user = server_user_field.get()
    server_pass = server_password_field.get()
    server.destroy()

def is_number(n):
    try:
        float(n)
        return True
    except:
        return False

def month_conversion(val):
    month_check = {
        'JANUARY': '1',
        'FEBRUARY': '2',
        'MARCH': '3',
        'APRIL': '4',
        'MAY': '5',
        'JUNE': '6',
        'JULY': '7',
        'AUGUST': '8',
        'SEPTEMBER': '9',
        'OCTOBER': '10',
        'NOVEMBER': '11',
        'DECEMBER': '12'
    }
    return month_check.get(val, 0)

def date_id(month, year):
    if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) and month == 2):
        return str(year) + "0229"
    else:
        return str(year) + [None, "0131", "0228", "0331", "0430", "0531", "0630", "0731", "0831", "0930", "1031", "1130", "1231"][int(month)]

# Main update function.
def update(payload={'product': 'CF6', 'station': 'NYC', 'recent': 'yes' }, check_assert=True):
    global ws
    global wb
    global conn
    global cursor
    # Load the NOAA website.
    try:
        source = requests.post(url='https://w2.weather.gov/climate/getclimate.php?wfo=okx', data=payload)
        report = BeautifulSoup(source.text, 'lxml').pre.text.splitlines()
    except:
        MessageBox(None,
                    'Connection to NOAA website failed. Perhaps the service is down, or the URL has changed. Try again later.', 
                    'Fatal Error', 0x10 | 0x1000)
        raise

    # Test workbook access.
    try:
        wb.save('NOAA_Weather.xlsx')
    except PermissionError:
        MessageBox(None, 'Workbook access denied. If the file is open, close it and retry.', 'Error', 0x10 | 0x1000)
        raise
    
    # Parse NOAA report and add weather data for all available days to complete_data.
    complete_data = []
    i = 19
    while ("=" not in report[i]):
        complete_data.append(report[i].split(" "))
        i += 1
            
    # Clean all whitespace from each date_data, change all 'T' to '0.001', remove redundant data.
    new_data = []
    date = ''
    try:
        for col in ws.iter_cols(min_row=ws.max_row, max_col=3):
            for cell in col:
                if int(cell.value) < 10:
                    date += '0' + str(int(cell.value))
                else:
                    date += str(int(cell.value))

        max_date = date[2:6] + date[0:2] + date[6:]
    except ValueError:
        MessageBox(None, 
                    'Excel worksheet formatted incorrectly. Delete blank rows below last visible data and try again.', 
                    'Critical Error', 0x10 | 0x1000)
        raise

    for date_data in complete_data:
        for i, x in enumerate(date_data):
            date_data[i] = 0.001 if x == 'T' else x
        while ('' in date_data):
            date_data.remove('')
        date_data.insert(0, month_conversion(report[7].split(" ").pop()))
        date_data.insert(1, report[8].split(" ").pop())
        if (len(date_data) == 20):
            date_data.insert(len(date_data) - 2, '')
        if len(date_data[0]) < 2:
            month = '0' + str(date_data[0])
        else:
            month = str(date_data[0])
        if len(date_data[2]) < 2:
            day = '0' + str(date_data[2])
        else:
            day = str(date_data[2])

        year = str(date_data[1])
        date = year + month + day

        if (date > max_date):
            new_data.append(date_data)
    
    # Check if existing data is already up to date or needs updating.
    if (new_data):
        if (check_assert):
            if (MessageBox(None, 'New data was found. Proceed with update?', 'Confirmation', 0x30 | 0x01 | 0x1000) == 2):
                raise KeyboardInterrupt
            # Detect if any missing data from previous months exist.
            assert new_data[0][0] == str(ws.cell(ws.max_row, 1).value) and new_data[0][1] == str(ws.cell(ws.max_row, 2).value)
    else:
        MessageBox(None, 'No new data was found. Update procedure aborted.', 'Error', 0x10 | 0x1000)
        raise KeyboardInterrupt
        
    # Update the worksheet.
    for row in new_data:
        ws.append([float(x) if is_number(x) else x for x in row])
    for row in ws.iter_rows(min_row=ws.max_row - len(new_data) + 1, max_col=24, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name='Arial', size=12)
            cell.alignment = Alignment(horizontal='center')
        
    # Update the SQL server.
    cursor.execute("USE //REDACTED//;")
    cursor.execute("TRUNCATE TABLE [weather].[NOAA_Weather];")
    for row in ws.iter_rows(min_row=2, max_col=23, values_only=True):
        while(len(row) < 23):
            row.append("NULL")
        try:
            cursor.execute("INSERT INTO [weather].[NOAA_Weather] VALUES "
                            + str(tuple(x if is_number(x) else "NULL" for x in row)).replace("'NULL'", "NULL") + ";")
        except:
            MessageBox(None, 'SQL query failed. Please try again later.', 'Fatal Error', 0x10 | 0x1000)
            raise

    # Commit updates.
    conn.commit()
    
    # Save the workbook.
    wb.save('NOAA_Weather.xlsx')

root = Tk()
root.title('NOAAUpdater')
root_status = Label(root, text='This program is password protected. Enter authorization below.', font=('Calibri', 11))
root_status.grid(row=0, column=0)
root_field = Entry(root, width=30, show='*')
root_field.focus()
root_field.grid(row=1, column=0)
root_button = Button(root, text='Authenticate', command=auth, bg='lime')
root_button.grid(row=2, column=0, pady=5)
root.lift()
root.attributes('-topmost', True)
root.resizable(width=False, height=False)
root.update()
root.mainloop()

# Load NOAA_Weather.xlsx to write to
try:
    wb = load_workbook(filename='NOAA_Weather.xlsx')
except FileNotFoundError:
    MessageBox(None, 
                'NOAA_Weather.xlsx was not found. Please place this script in the same directory as the file.', 
                'Error', 0x10 | 0x1000)
    raise
ws = wb["Weather Data"]

if (cleared):
    # Ask user for server credentials.
    server = Tk()
    server.title('NOAAUpdater')
    server_status = Label(server, text='Enter SQL Server //REDACTED// credentials below.', font=('Calibri', 11))
    server_status.grid(row=0,columnspan=2)
    server_user_text = Label(server, text='Username')
    server_user_text.grid(row=1)
    server_password_text = Label(server, text='Password')
    server_password_text.grid(row=2)
    server_user_field = Entry(server, width=35)
    server_user_field.focus()
    server_user_field.grid(row=1, column=1)
    server_password_field = Entry(server, show='*', width=35)
    server_password_field.grid(row=2, column=1)
    server_button = Button(server, text='Submit', command=server_submit, bg='lime')
    server_button.grid(row=3, columnspan=2, pady=5)
    server.lift()
    server.attributes('-topmost', True)
    server.resizable(width=False, height=False)
    server.update()
    server.mainloop()

    # Connect to the SQL server.
    try:
        conn = pyodbc.connect("DRIVER={{SQL Server Native Client 11.0}}; SERVER=//REDACTED//; DATABASE=//REDACTED//; UID={0}; PWD={1};".format(server_user, server_pass))
        cursor = conn.cursor()
    except:
        MessageBox(None, 'Connection to SQL server failed. Perhaps the server is down, or the server credentials were changed.', 
                    'Fatal Error', 0x10 | 0x1000)
        raise KeyboardInterrupt

    try:
        update()
    except AssertionError:
        # Add data from previous months.
        current_month = ws.cell(ws.max_row, 1).value
        current_year = ws.cell(ws.max_row, 2).value
        while (date_id(current_month, current_year) < date_id(datetime.now().month, datetime.now().year)):
            update(payload={'product': 'CF6', 'station': 'NYC', 'recent': 'no', 'date': date_id(current_month, current_year)}, check_assert=False)
            current_month += 1
            if (current_month > 12):
                current_month = 1
                current_year += 1
        if (datetime.now().day != 1):
            update(check_assert=False)

    # Close the SQL connection.
    cursor.close()
    conn.close()
    
    # Terminate program.
    MessageBox(None, 'Update to Excel file and //REDACTED// SQL database completed. You may verify the new data to confirm accuracy.',
                'Complete', 0x40 | 0x1000)

# Created by Patrick Fan.